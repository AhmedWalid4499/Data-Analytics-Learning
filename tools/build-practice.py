#!/usr/bin/env python3
"""Builds themed practice datasets (xlsx + csv) and the practice hub + detail pages."""
import os, csv, random, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = "/home/claude/dpm-learning-hub"
PDIR = os.path.join(ROOT, "assets", "practice")
os.makedirs(PDIR, exist_ok=True)

def L(i): return chr(65 + i)  # 0->A
def money0(v): return "$" + format(round(v), ",")
def money2(v): return "$" + format(round(v, 2), ",.2f")
def num(v, d=0): return format(round(v, d), "," + ("." + str(d) + "f" if d else ""))
def pct(v, d=1): return format(v * 100, "." + str(d) + "f") + "%"

# ----------------------------------------------------------------- datasets
def gen_coffee():
    random.seed(11)
    products = {"Latte":("Coffee",4.50),"Cappuccino":("Coffee",4.20),"Espresso":("Coffee",3.00),
        "Mocha":("Coffee",4.80),"Americano":("Coffee",3.50),"Green Tea":("Tea",3.20),
        "Chai Latte":("Tea",4.00),"Croissant":("Food",3.50),"Muffin":("Food",2.80),
        "Bagel":("Food",3.20),"Sandwich":("Food",6.50),"Iced Coffee":("Cold",4.50),"Smoothie":("Cold",5.50)}
    stores=["Downtown","Airport","University","Mall"]; pays=["Card","Cash","Mobile"]
    pw=[0.55,0.25,0.20]; rows=[]; start=datetime.date(2025,1,1)
    names=list(products.keys())
    for i in range(220):
        p=random.choices(names, weights=[5,4,3,3,4,2,2,3,3,2,2,4,2])[0]
        cat,price=products[p]; q=random.choices([1,2,3],weights=[7,2,1])[0]
        d=start+datetime.timedelta(days=random.randint(0,89))
        rows.append({"Transaction ID":"T"+str(1001+i),"Date":d.isoformat(),"Store":random.choice(stores),
            "Product":p,"Category":cat,"Quantity":q,"Unit Price":price,"Total":round(q*price,2),
            "Payment":random.choices(pays,weights=pw)[0]})
    return rows

def gen_hr():
    random.seed(22)
    depts={"Engineering":(70000,120000),"Sales":(45000,90000),"Marketing":(48000,85000),
        "Finance":(55000,100000),"HR":(45000,80000),"Support":(35000,60000)}
    titles={"Engineering":["Engineer","Senior Engineer","Lead Engineer"],"Sales":["Sales Rep","Account Manager","Sales Lead"],
        "Marketing":["Marketing Exec","Content Lead","Brand Manager"],"Finance":["Analyst","Senior Analyst","Controller"],
        "HR":["HR Officer","Recruiter","HR Manager"],"Support":["Support Agent","Senior Agent","Support Lead"]}
    locs=["London","Berlin","Madrid","Paris","Remote"]; first=["Alex","Sam","Maria","John","Lena","Omar","Priya","Tom","Sara","Nina","Luca","Eva","Max","Yara","Ravi","Anna","Leo","Mia","Jon","Zoe"]
    last=["Smith","Khan","Garcia","Müller","Rossi","Dubois","Patel","Brown","Nowak","Silva","Ivanov","Costa","Wagner","Haddad","Reyes","Lee","Novak","Ali","Petrov","Marin"]
    rows=[]
    for i in range(150):
        dep=random.choice(list(depts)); lo,hi=depts[dep]
        hire=datetime.date(random.randint(2015,2024),random.randint(1,12),random.randint(1,28))
        rows.append({"Employee ID":"E"+str(2001+i),"Name":random.choice(first)+" "+random.choice(last),
            "Department":dep,"Job Title":random.choice(titles[dep]),"Location":random.choice(locs),
            "Gender":random.choice(["F","M"]),"Hire Date":hire.isoformat(),
            "Salary":int(round(random.uniform(lo,hi)/500)*500),"Rating":random.choice([2,3,3,4,4,4,5]),
            "Status":random.choices(["Active","Left"],weights=[0.84,0.16])[0]})
    return rows

def gen_ecom():
    random.seed(33)
    cats={"Furniture":["Chairs","Tables","Bookcases","Furnishings"],
        "Office Supplies":["Binders","Paper","Storage","Art","Labels"],
        "Technology":["Phones","Accessories","Machines","Copiers"]}
    segs=["Consumer","Corporate","Home Office"]; countries=["USA","UK","Germany","France","Canada","Australia"]
    custs=["Acme Co","Globex","Initech","Umbrella","Soylent","Stark Ind","Wayne LLC","Hooli","Vandelay","Wonka","Pied Piper","Gekko Cap"]
    rows=[]; start=datetime.date(2024,1,1)
    for i in range(250):
        cat=random.choice(list(cats)); sub=random.choice(cats[cat])
        qty=random.randint(1,8); base={"Furniture":250,"Office Supplies":40,"Technology":300}[cat]
        sales=round(base*qty*random.uniform(0.5,2.2),2); disc=random.choice([0,0,0,0.1,0.2,0.3])
        margin=random.uniform(-0.15,0.35)-disc*0.4
        d=start+datetime.timedelta(days=random.randint(0,364))
        rows.append({"Order ID":"O"+str(5001+i),"Order Date":d.isoformat(),"Customer":random.choice(custs),
            "Segment":random.choice(segs),"Country":random.choice(countries),"Category":cat,"Sub-Category":sub,
            "Sales":sales,"Quantity":qty,"Discount":disc,"Profit":round(sales*margin,2)})
    return rows

def gen_movies():
    random.seed(44)
    genres=["Action","Drama","Comedy","Sci-Fi","Horror","Animation","Thriller"]
    dirs=["A. Reed","B. Cole","C. Vance","D. Park","E. Ford","F. Nair","G. Stone","H. Liu","I. Roth","J. Mara"]
    adj=["Last","Dark","Lost","Final","Hidden","Broken","Silent","Golden","Rising","Frozen","Wild","Secret",
         "Endless","Crimson","Distant","Burning","Quiet","Northern","Electric","Velvet"]
    noun=["Horizon","Empire","Echo","Legacy","Shadow","Voyage","Promise","Kingdom","Signal","Garden","Storm","Mirror",
          "Harbour","Compass","Lantern","Circuit","Meridian","Anthem","Cascade","Ember"]
    rows=[]; used=set()
    for i in range(180):
        attempts=0; t=None
        while True:
            cand=random.choice(adj)+" "+random.choice(noun)
            attempts+=1
            if cand not in used:
                t=cand; break
            if attempts>40:
                t=cand+" "+random.choice(["II","Reborn","Returns","Dawn"])
                if t not in used: break
        used.add(t)
        g=random.choice(genres); budget=round(random.uniform(5,250),1)
        box=round(budget*random.uniform(0.3,5.0),1); rating=round(random.uniform(4.5,8.9),1)
        rows.append({"Title":t,"Genre":g,"Year":random.randint(2005,2023),"Runtime":random.randint(88,168),
            "Budget ($M)":budget,"Box Office ($M)":box,"IMDb Rating":rating,"Director":random.choice(dirs)})
    return rows

# ----------------------------------------------------------------- write files
def write_files(slug, rows):
    headers=list(rows[0].keys())
    # csv
    with open(os.path.join(PDIR, slug+".csv"),"w",newline="",encoding="utf-8") as f:
        w=csv.DictWriter(f,fieldnames=headers); w.writeheader(); w.writerows(rows)
    # xlsx
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Data"
    hf=Font(bold=True,color="FFFFFF"); fill=PatternFill("solid",fgColor="1F4E79")
    ws.append(headers)
    for c in ws[1]: c.font=hf; c.fill=fill; c.alignment=Alignment(horizontal="left")
    for r in rows: ws.append([r[h] for h in headers])
    ws.freeze_panes="A2"
    for i,h in enumerate(headers):
        width=max(len(str(h)), max((len(str(r[h])) for r in rows), default=8))+2
        ws.column_dimensions[L(i)].width=min(width,28)
    ws.auto_filter.ref="A1:"+L(len(headers)-1)+str(len(rows)+1)
    wb.save(os.path.join(PDIR, slug+".xlsx"))
    return headers

# ----------------------------------------------------------------- HTML helpers
NAV = ('<nav class="nav"><div class="wrap">\n'
 '  <a class="brand" href="index.html"><span class="dot"></span> DPM Learning Hub</a>\n'
 '  <button class="menu-btn" aria-label="Menu">\u2630</button>\n'
 '  <div class="nav-links">\n'
 '    <a href="index.html">Home</a>\n'
 '    <a href="data-foundations.html">Foundations</a>\n'
 '    <a href="excel.html">Excel</a>\n'
 '    <a href="sql.html">SQL</a>\n'
 '    <a href="powerbi.html">Power BI</a>\n'
 '    <a href="powerautomate.html">Power Automate</a>\n'
 '    <a href="courses.html">Courses</a>\n'
 '    <a href="downloads.html">Downloads</a>\n'
 '    <a href="practice.html" aria-current="page">Practice</a>\n'
 '  </div>\n</div></nav>\n')

def head(title):
    return ('<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">\n'
    '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
    '<title>'+title+' \u00b7 DPM Learning Hub</title>\n'
    '<link rel="preconnect" href="https://fonts.googleapis.com">\n'
    '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>\n'
    '<link href="https://fonts.googleapis.com/css2?family=Bricolage+Grotesque:opsz,wght@12..96,600;12..96,700;12..96,800&family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@500;600;700&display=swap" rel="stylesheet">\n'
    '<link rel="stylesheet" href="assets/css/styles.css">\n</head>\n')

def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

# ----------------------------------------------------------------- themes
def C(i): return L(i)  # column letter

def build_theme(slug, name, emoji, color, blurb, skills, rows, columns_desc, tasks_fn, last):
    headers=write_files(slug, rows)
    n=len(rows); lastrow=n+1
    # column letters map
    idx={h:L(i) for i,h in enumerate(headers)}
    tasks=tasks_fn(rows, idx, lastrow)

    # ---- detail page ----
    h=head("Practice \u2014 "+name)
    h+='<body data-theme="hub" style="--accent:'+color+'">\n<div class="scrollbar"></div>\n'+NAV
    h+=('<header class="hero" style="padding-bottom:14px"><div class="wrap">\n'
        '  <span class="tool-badge"><span class="glyph">'+emoji+'</span> Practice dataset</span>\n'
        '  <h1 style="font-size:2.3rem">'+name+'</h1>\n'
        '  <p class="lead">'+blurb+'</p>\n'
        '  <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:14px">\n'
        '    <a class="btn" href="assets/practice/'+slug+'.xlsx" download>\u2b07 Download Excel (.xlsx)</a>\n'
        '    <a class="btn ghost" href="assets/practice/'+slug+'.csv" download>\u2b07 CSV</a>\n'
        '    <a class="btn ghost" href="practice.html">\u2190 All themes</a>\n'
        '  </div>\n'
        '  <p class="hint" style="color:var(--muted);font-size:.9rem;margin-top:12px">'+str(n)+' rows \u00b7 '+str(len(headers))+' columns \u00b7 fictional data for practice. Works in Excel, Google Sheets or Power BI.</p>\n'
        '</div></header>\n')
    # columns
    h+='<section><div class="wrap">\n  <span class="eyebrow">What\'s in the file</span>\n  <h2 style="font-size:1.7rem;margin:.2em 0 .5em">Columns</h2>\n  <table class="tbl"><thead><tr><th>Column</th><th>What it holds</th></tr></thead><tbody>\n'
    for cn,desc in columns_desc:
        h+='    <tr><td><code>'+esc(cn)+'</code></td><td>'+desc+'</td></tr>\n'
    h+='  </tbody></table>\n'
    # preview
    h+='  <h3 style="margin:26px 0 8px">First few rows</h3>\n  <div style="overflow:auto"><table class="tbl"><thead><tr>'
    for hd in headers: h+='<th>'+esc(hd)+'</th>'
    h+='</tr></thead><tbody>\n'
    for r in rows[:6]:
        h+='    <tr>'+''.join('<td>'+esc(r[hd])+'</td>' for hd in headers)+'</tr>\n'
    h+='  </tbody></table></div>\n</div></section>\n'
    # tasks
    h+='<section style="background:var(--surface-2)"><div class="wrap">\n  <span class="eyebrow">Practice tasks</span>\n  <h2 style="font-size:1.7rem;margin:.2em 0 .3em">'+str(len(tasks))+' tasks \u2014 try them, then reveal the answer</h2>\n  <p class="lead">Build each one yourself in Excel (or Power BI) first. The answer shows a working Excel formula, the result from this dataset, and how you\'d do it in Power BI.</p>\n'
    for k,t in enumerate(tasks,1):
        h+=('  <div class="card" style="margin:12px 0">\n'
            '    <p style="margin:.1em 0 .5em"><b>Task '+str(k)+'.</b> '+t["q"]+'</p>\n'
            '    <details class="reveal"><summary>Show answer</summary><div class="ans">\n'
            '      <div class="code"><button class="copy">Copy</button><span class="code-text">'+esc(t["formula"])+'</span></div>\n'
            '      <p><b>Answer:</b> '+t["ans"]+'</p>\n'
            '      <p style="color:var(--ink-soft)"><b>In Power BI:</b> '+t["pbi"]+'</p>\n'
            '    </div></details>\n  </div>\n')
    h+='</div></section>\n'
    h+=('<footer class="foot"><div class="wrap"><p>Practice \u00b7 '+name+' \u00b7 DPM Learning Hub. <a href="practice.html">\u2190 Back to all practice themes</a></p></div></footer>\n'
        '<script src="assets/js/main.js"></script>\n</body>\n</html>\n')
    open(os.path.join(ROOT,"practice-"+last+".html"),"w",encoding="utf-8").write(h)
    return {"slug":slug,"name":name,"emoji":emoji,"color":color,"blurb":blurb,"skills":skills,
            "rows":n,"cols":len(headers),"tasks":len(tasks),"page":"practice-"+last+".html"}

# ----------------------------------------------------------------- task builders
def t_coffee(rows, x, last):
    tot=sum(r["Total"] for r in rows); cnt=len(rows)
    by_store={}; by_prod={}
    for r in rows:
        by_store[r["Store"]]=by_store.get(r["Store"],0)+r["Total"]
        by_prod[r["Product"]]=by_prod.get(r["Product"],0)+r["Total"]
    latte=sum(r["Quantity"] for r in rows if r["Product"]=="Latte")
    over20=sum(1 for r in rows if r["Total"]>20)
    airport=sum(r["Total"] for r in rows if r["Store"]=="Airport")
    cardp=sum(1 for r in rows if r["Payment"]=="Card")/cnt
    top_store=max(by_store,key=by_store.get); top_prod=max(by_prod,key=by_prod.get)
    return [
     {"q":"What is the <b>total revenue</b> across all transactions?","formula":"=SUM(H2:H"+str(last)+")","ans":money2(tot),"pbi":"A Card visual showing SUM(Total)."},
     {"q":"What is the <b>average transaction value</b>?","formula":"=AVERAGE(H2:H"+str(last)+")","ans":money2(sum(r['Total'] for r in rows)/len(rows)),"pbi":"A Card with AVERAGE(Total)."},
     {"q":"How many <b>Lattes</b> were sold (total quantity)?","formula":'=SUMIF(D:D,"Latte",F:F)',"ans":num(latte),"pbi":"Card: CALCULATE(SUM(Quantity), Product = \"Latte\")."},
     {"q":"How many transactions were <b>over $20</b>?","formula":'=COUNTIF(H:H,">20")',"ans":num(over20),"pbi":"Card with a measure: COUNTROWS(FILTER(Sales, Sales[Total]>20))."},
     {"q":"What is the total revenue from the <b>Airport</b> store?","formula":'=SUMIF(C:C,"Airport",H:H)',"ans":money2(airport),"pbi":"Filter a SUM(Total) measure to Store = Airport."},
     {"q":"Which <b>store</b> has the highest total revenue?","formula":'=SUMIFS(H:H,C:C,"Downtown")  \u2026 compare per store (or use a PivotTable: Store \u2192 Rows, Sum of Total \u2192 Values)',"ans":"<b>"+top_store+"</b> ("+money2(by_store[top_store])+")","pbi":"Bar chart of SUM(Total) by Store, sorted descending."},
     {"q":"Which <b>product</b> generates the most revenue?","formula":'PivotTable: Product \u2192 Rows, Sum of Total \u2192 Values, sort descending (or SUMIF per product)',"ans":"<b>"+top_prod+"</b> ("+money2(by_prod[top_prod])+")","pbi":"Bar chart of SUM(Total) by Product, Top N = 1."},
     {"q":"What <b>percentage</b> of transactions were paid by <b>Card</b>?","formula":'=COUNTIF(I:I,"Card")/COUNTA(I2:I'+str(last)+')',"ans":pct(cardp),"pbi":"Card visual: DIVIDE(COUNTROWS(FILTER(Sales,Payment=\"Card\")), COUNTROWS(Sales))."},
    ]

def t_hr(rows, x, last):
    n=len(rows); sal=[r["Salary"] for r in rows]
    active=sum(1 for r in rows if r["Status"]=="Active"); left=n-active
    eng=sum(r["Salary"] for r in rows if r["Department"]=="Engineering")
    sales=[r["Salary"] for r in rows if r["Department"]=="Sales"]
    by_dep={}
    for r in rows: by_dep.setdefault(r["Department"],[]).append(r["Salary"])
    avg_dep={d:sum(v)/len(v) for d,v in by_dep.items()}; top=max(avg_dep,key=avg_dep.get)
    rating=sum(r["Rating"] for r in rows)/n
    return [
     {"q":"What is the <b>total headcount</b> in the file?","formula":"=COUNTA(A2:A"+str(last)+")","ans":num(n),"pbi":"Card: COUNTROWS(Employees)."},
     {"q":"How many employees are currently <b>Active</b>?","formula":'=COUNTIF(J:J,"Active")',"ans":num(active),"pbi":"Card filtered to Status = Active."},
     {"q":"What is the <b>average salary</b> across everyone?","formula":"=AVERAGE(H2:H"+str(last)+")","ans":money0(sum(sal)/n),"pbi":"Card: AVERAGE(Salary)."},
     {"q":"What is the <b>total payroll</b> of the <b>Engineering</b> department?","formula":'=SUMIF(C:C,"Engineering",H:H)',"ans":money0(eng),"pbi":"SUM(Salary) filtered to Department = Engineering."},
     {"q":"What is the <b>average salary in Sales</b>?","formula":'=AVERAGEIF(C:C,"Sales",H:H)',"ans":money0(sum(sales)/len(sales)),"pbi":"AVERAGE(Salary) with Department = Sales."},
     {"q":"What is the <b>highest salary</b> in the company?","formula":"=MAX(H2:H"+str(last)+")","ans":money0(max(sal)),"pbi":"Card: MAX(Salary)."},
     {"q":"Which <b>department</b> has the highest <b>average salary</b>?","formula":'PivotTable: Department \u2192 Rows, Average of Salary \u2192 Values (or AVERAGEIF per dept)',"ans":"<b>"+top+"</b> ("+money0(avg_dep[top])+")","pbi":"Bar chart of AVERAGE(Salary) by Department."},
     {"q":"How many employees have <b>Left</b> (attrition count)?","formula":'=COUNTIF(J:J,"Left")',"ans":num(left)+"  ("+pct(left/n)+" of staff)","pbi":"Card filtered to Status = Left; show as % with DIVIDE."},
     {"q":"What is the <b>average performance rating</b>?","formula":"=AVERAGE(I2:I"+str(last)+")","ans":num(rating,2),"pbi":"Card: AVERAGE(Rating)."},
    ]

def t_ecom(rows, x, last):
    tot=sum(r["Sales"] for r in rows); prof=sum(r["Profit"] for r in rows)
    tech=sum(r["Sales"] for r in rows if r["Category"]=="Technology")
    loss=sum(1 for r in rows if r["Profit"]<0)
    disc=sum(r["Discount"] for r in rows)/len(rows)
    germ=sum(r["Sales"] for r in rows if r["Country"]=="Germany")
    by_cat={}
    for r in rows: by_cat[r["Category"]]=by_cat.get(r["Category"],0)+r["Profit"]
    topcat=max(by_cat,key=by_cat.get)
    cons_tech=sum(r["Sales"] for r in rows if r["Segment"]=="Consumer" and r["Category"]=="Technology")
    return [
     {"q":"What are <b>total sales</b>?","formula":"=SUM(H2:H"+str(last)+")","ans":money0(tot),"pbi":"Card: SUM(Sales)."},
     {"q":"What is <b>total profit</b>?","formula":"=SUM(K2:K"+str(last)+")","ans":money0(prof),"pbi":"Card: SUM(Profit)."},
     {"q":"What is the overall <b>profit margin</b> (profit \u00f7 sales)?","formula":"=SUM(K2:K"+str(last)+")/SUM(H2:H"+str(last)+")","ans":pct(prof/tot),"pbi":"Measure: DIVIDE(SUM(Profit), SUM(Sales)), formatted as %."},
     {"q":"What are total sales in the <b>Technology</b> category?","formula":'=SUMIF(F:F,"Technology",H:H)',"ans":money0(tech),"pbi":"SUM(Sales) filtered to Category = Technology."},
     {"q":"How many orders are <b>loss-making</b> (Profit &lt; 0)?","formula":'=COUNTIF(K:K,"<0")',"ans":num(loss),"pbi":"Card: COUNTROWS(FILTER(Orders, Orders[Profit]<0))."},
     {"q":"Which <b>category</b> is the most profitable?","formula":'PivotTable: Category \u2192 Rows, Sum of Profit \u2192 Values (or SUMIF per category)',"ans":"<b>"+topcat+"</b> ("+money0(by_cat[topcat])+")","pbi":"Bar chart of SUM(Profit) by Category."},
     {"q":"What is the <b>average discount</b> given?","formula":"=AVERAGE(J2:J"+str(last)+")","ans":pct(disc),"pbi":"Card: AVERAGE(Discount) as %."},
     {"q":"What are total sales in <b>Germany</b>?","formula":'=SUMIF(E:E,"Germany",H:H)',"ans":money0(germ),"pbi":"SUM(Sales) filtered to Country = Germany; or a Map visual."},
     {"q":"What are total sales for the <b>Consumer</b> segment in <b>Technology</b>?","formula":'=SUMIFS(H:H,D:D,"Consumer",F:F,"Technology")',"ans":money0(cons_tech),"pbi":"SUM(Sales) with two filters: Segment = Consumer, Category = Technology."},
    ]

def t_movies(rows, x, last):
    box=[r["Box Office ($M)"] for r in rows]; rate=[r["IMDb Rating"] for r in rows]
    action=sum(1 for r in rows if r["Genre"]=="Action")
    scifi=[r["Budget ($M)"] for r in rows if r["Genre"]=="Sci-Fi"]
    over8=sum(1 for r in rows if r["IMDb Rating"]>8)
    box2019=sum(r["Box Office ($M)"] for r in rows if r["Year"]==2019)
    by_gen={}
    for r in rows: by_gen.setdefault(r["Genre"],[]).append(r["IMDb Rating"])
    avg_gen={g:sum(v)/len(v) for g,v in by_gen.items()}; topg=max(avg_gen,key=avg_gen.get)
    by_dir={}
    for r in rows: by_dir[r["Director"]]=by_dir.get(r["Director"],0)+r["Box Office ($M)"]
    topd=max(by_dir,key=by_dir.get)
    return [
     {"q":"What is the <b>highest box office</b> figure ($M)?","formula":"=MAX(F2:F"+str(last)+")","ans":"$"+num(max(box),1)+"M","pbi":"Card: MAX(Box Office)."},
     {"q":"What is the <b>average IMDb rating</b>?","formula":"=AVERAGE(G2:G"+str(last)+")","ans":num(sum(rate)/len(rate),2),"pbi":"Card: AVERAGE(IMDb Rating)."},
     {"q":"How many <b>Action</b> movies are there?","formula":'=COUNTIF(B:B,"Action")',"ans":num(action),"pbi":"Card filtered to Genre = Action."},
     {"q":"What is the <b>average budget</b> of <b>Sci-Fi</b> movies ($M)?","formula":'=AVERAGEIF(B:B,"Sci-Fi",E:E)',"ans":"$"+num(sum(scifi)/len(scifi),1)+"M","pbi":"AVERAGE(Budget) filtered to Genre = Sci-Fi."},
     {"q":"How many movies are rated <b>above 8.0</b>?","formula":'=COUNTIF(G:G,">8")',"ans":num(over8),"pbi":"Card: COUNTROWS(FILTER(Movies, Rating>8))."},
     {"q":"What is the <b>total box office for 2019</b> ($M)?","formula":'=SUMIF(C:C,2019,F:F)',"ans":"$"+num(box2019,1)+"M","pbi":"SUM(Box Office) filtered to Year = 2019."},
     {"q":"Which <b>genre</b> has the highest average rating?","formula":'PivotTable: Genre \u2192 Rows, Average of IMDb Rating \u2192 Values','ans':"<b>"+topg+"</b> ("+num(avg_gen[topg],2)+")","pbi":"Bar chart of AVERAGE(Rating) by Genre."},
     {"q":"Which <b>director</b> has the highest total box office?","formula":'PivotTable: Director \u2192 Rows, Sum of Box Office \u2192 Values, sort descending',"ans":"<b>"+topd+"</b> ($"+num(by_dir[topd],1)+"M)","pbi":"Bar chart of SUM(Box Office) by Director, Top N = 1."},
    ]

# ----------------------------------------------------------------- build all
themes=[]
themes.append(build_theme("coffee-shop-sales","Coffee Shop Sales","\u2615","#B45309",
    "Point-of-sale transactions from a small coffee-shop chain \u2014 perfect for totals, conditional sums, averages and a first sales dashboard.",
    ["SUMIFS","COUNTIF","PivotTables","Charts"], gen_coffee(),
    [("Transaction ID","Unique id for each sale"),("Date","Date of the transaction"),("Store","Which branch (Downtown, Airport, University, Mall)"),
     ("Product","Item sold"),("Category","Coffee, Tea, Food or Cold"),("Quantity","Units sold"),("Unit Price","Price per unit ($)"),
     ("Total","Quantity \u00d7 Unit Price ($)"),("Payment","Card, Cash or Mobile")], t_coffee, "coffee"))

themes.append(build_theme("hr-employees","HR \u2014 Employees","\U0001F465","#6D28D9",
    "A company headcount file \u2014 great for lookups, averages by group, headcount and attrition, and an HR dashboard.",
    ["AVERAGEIFS","COUNTIF","XLOOKUP","KPIs"], gen_hr(),
    [("Employee ID","Unique employee id"),("Name","Employee name"),("Department","Team they belong to"),("Job Title","Role"),
     ("Location","Office or Remote"),("Gender","F or M"),("Hire Date","Date they joined"),("Salary","Annual salary ($)"),
     ("Rating","Performance rating 1\u20135"),("Status","Active or Left")], t_hr, "hr"))

themes.append(build_theme("ecommerce-orders","E-commerce Orders","\U0001F6D2","#0E7490",
    "Online retail orders with sales, discount and profit \u2014 ideal for margin analysis, profit by category and a Superstore-style dashboard.",
    ["SUMIFS","margins","FILTER","Maps"], gen_ecom(),
    [("Order ID","Unique order id"),("Order Date","Date the order was placed"),("Customer","Buying company"),("Segment","Consumer, Corporate, Home Office"),
     ("Country","Ship-to country"),("Category","Furniture, Office Supplies, Technology"),("Sub-Category","Finer product group"),
     ("Sales","Order value ($)"),("Quantity","Units ordered"),("Discount","Discount as a fraction (0\u20130.3)"),("Profit","Profit ($, can be negative)")], t_ecom, "ecommerce"))

themes.append(build_theme("movies","Movies","\U0001F3AC","#BE123C",
    "A film catalogue with budgets, box office and ratings \u2014 good for ranking, ratios (ROI), filtering and comparison charts.",
    ["MAX/MIN","AVERAGEIF","RANK","ratios"], gen_movies(),
    [("Title","Film title"),("Genre","Action, Drama, Comedy, Sci-Fi, Horror, Animation, Thriller"),("Year","Release year"),
     ("Runtime","Length in minutes"),("Budget ($M)","Production budget, $ millions"),("Box Office ($M)","Worldwide gross, $ millions"),
     ("IMDb Rating","Audience rating 0\u201310"),("Director","Director name")], t_movies, "movies"))

# ----------------------------------------------------------------- hub page
h=head("Practice")
h+='<body data-theme="hub">\n<div class="scrollbar"></div>\n'+NAV
h+=('<header class="hero"><div class="wrap">\n'
    '  <span class="tool-badge"><span class="glyph">\U0001F4CA</span> Practice library</span>\n'
    '  <h1>Practice on real-ish<br>data \u2014 your pick of theme.</h1>\n'
    '  <p class="lead">The lesson tracks use one delivery dataset, but skills stick when you practise on <strong>different</strong> data. Each theme below is a self-contained workbook with its own tasks and worked answers \u2014 download the sheet, try the tasks in Excel or Power BI, then reveal the solution. None of it is DPM-specific, so use whatever interests you.</p>\n'
    '</div></header>\n')
h+='<section><div class="wrap">\n  <span class="eyebrow">Choose a theme</span>\n  <h2 style="font-size:1.8rem;margin:.2em 0 .6em">'+str(len(themes))+' practice sets</h2>\n  <div class="grid cols-2">\n'
for t in themes:
    pills="".join('<span class="term">'+esc(s)+'</span>' for s in t["skills"])
    h+=('    <div class="card" style="border-top:4px solid '+t["color"]+'">\n'
        '      <div style="font-size:2rem;line-height:1">'+t["emoji"]+'</div>\n'
        '      <h3 style="margin:.3em 0 .2em;font-size:1.25rem">'+t["name"]+'</h3>\n'
        '      <p style="color:var(--ink-soft);margin:0 0 .6em">'+t["blurb"]+'</p>\n'
        '      <div class="pills" style="margin-bottom:.7em">'+pills+'</div>\n'
        '      <p class="hint" style="color:var(--muted);font-size:.85rem;margin:0 0 .8em">'+str(t["rows"])+' rows \u00b7 '+str(t["cols"])+' columns \u00b7 '+str(t["tasks"])+' tasks</p>\n'
        '      <a class="btn" href="'+t["page"]+'" style="--accent:'+t["color"]+'">Open theme \u2192</a>\n'
        '    </div>\n')
h+='  </div>\n'
h+=('  <div class="note" style="margin-top:24px"><span class="ic">i</span><div>Every task lists a working Excel formula <em>and</em> a Power BI approach, plus the exact answer from the data so you can check yourself. More themes will be added over time.</div></div>\n')
h+='</div></section>\n'
h+=('<footer class="foot"><div class="wrap"><p>Practice library \u00b7 DPM Learning Hub. <a href="index.html">\u2190 Back to the hub</a></p></div></footer>\n'
    '<script src="assets/js/main.js"></script>\n</body>\n</html>\n')
open(os.path.join(ROOT,"practice.html"),"w",encoding="utf-8").write(h)

print("Built", len(themes), "themes:")
for t in themes: print("  -", t["name"], "->", t["page"], "(", t["rows"], "rows,", t["tasks"], "tasks )")
print("Datasets in assets/practice/:", sorted(os.listdir(PDIR)))
